"""
📚 ALGEBRA LINEAL - Módulo Core
==============================

Funciones principales para trabajar con álgebra lineal usando
Google Sheets (por nombre o enlace) o archivos Excel locales.
Este módulo contiene toda la lógica del sistema.

Instalado desde PyPI: pip install algebra-lineal-sheets
"""

import os
import re
import inspect

import gspread
from gspread.utils import ValueRenderOption
from google.auth import default
import numpy as np

# Variables globales del paquete
gc = None                      # Cliente de Google Sheets (None hasta configurar())
spreadsheet_name = 'matrices'  # Fuente activa: nombre, enlace o ruta .xlsx


# ============================================================
# Detección de la fuente de datos
# ============================================================

def _es_enlace(fuente):
    """True si la fuente es un enlace (URL) de Google Sheets."""
    return isinstance(fuente, str) and fuente.lower().startswith('http')


def _es_excel(fuente):
    """True si la fuente es una ruta a un archivo Excel local (.xlsx)."""
    return isinstance(fuente, str) and fuente.lower().endswith('.xlsx')


# ============================================================
# Backends: misma interfaz para Google Sheets y Excel local
# ============================================================

class _BackendGoogle:
    """Lee y escribe matrices en un Google Sheet (por nombre o enlace)."""

    def __init__(self, fuente):
        if _es_enlace(fuente):
            self._ss = gc.open_by_url(fuente)
        else:
            self._ss = gc.open(fuente)
        self.descripcion = f"☁️ Google Sheets: '{self._ss.title}'"
        self._pestanas = {ws.title: ws for ws in self._ss.worksheets()}

    def nombres_pestanas(self):
        return list(self._pestanas.keys())

    def leer(self, nombre):
        # UNFORMATTED: los números llegan como números, no como texto formateado
        # por la configuración regional (evita que "3,5" se lea como texto).
        return self._pestanas[nombre].get_all_values(
            value_render_option=ValueRenderOption.unformatted)

    def escribir(self, nombre, datos, sobrescribir=True):
        """Devuelve 'creada'/'actualizada', o None si existe y sobrescribir=False."""
        filas_datos = len(datos)
        cols_datos = max((len(fila) for fila in datos), default=1)

        if nombre in self._pestanas:
            if not sobrescribir:
                return None
            worksheet = self._pestanas[nombre]
            worksheet.clear()
            if (filas_datos > worksheet.row_count or
                    cols_datos > worksheet.col_count):
                worksheet.resize(rows=max(filas_datos, worksheet.row_count),
                                 cols=max(cols_datos, worksheet.col_count))
            accion = "actualizada"
        else:
            worksheet = self._ss.add_worksheet(
                title=nombre,
                rows=max(50, filas_datos),
                cols=max(20, cols_datos))
            self._pestanas[nombre] = worksheet
            accion = "creada"

        if datos:
            worksheet.update(values=datos, range_name='A1')
        return accion


class _BackendExcel:
    """Lee y escribe matrices en un archivo Excel local (.xlsx)."""

    def __init__(self, ruta, crear_si_no_existe=False):
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Falta openpyxl para el modo Excel. Ejecuta: pip install openpyxl")
        self._openpyxl = openpyxl
        self._ruta = ruta
        self._nuevo = not os.path.exists(ruta)
        self._aviso_formulas_dado = False
        if self._nuevo and not crear_si_no_existe:
            raise FileNotFoundError(f"No existe el archivo: {ruta}")
        self.descripcion = f"📊 Excel local: {os.path.basename(ruta)}"

    def _cargar(self, para_escribir=False):
        # Para leer: data_only=True devuelve los valores calculados de las fórmulas.
        # Para escribir: sin data_only, para no perder fórmulas de otras pestañas.
        return self._openpyxl.load_workbook(self._ruta, data_only=not para_escribir)

    def nombres_pestanas(self):
        if self._nuevo:
            return []
        return self._cargar().sheetnames

    def leer(self, nombre):
        # Valores nativos de openpyxl (float/int/None): sin pasar por texto,
        # el separador decimal regional nunca interfiere.
        hoja = self._cargar()[nombre]
        return [list(fila) for fila in hoja.iter_rows(values_only=True)]

    @staticmethod
    def _contiene_formulas(libro):
        for hoja in libro.worksheets:
            for fila in hoja.iter_rows(values_only=True):
                for valor in fila:
                    if isinstance(valor, str) and valor.startswith('='):
                        return True
        return False

    def escribir(self, nombre, datos, sobrescribir=True):
        """Devuelve 'creada'/'actualizada', o None si existe y sobrescribir=False."""
        if self._nuevo:
            libro = self._openpyxl.Workbook()
            hoja = libro.active
            hoja.title = nombre
            accion = "creada"
        else:
            libro = self._cargar(para_escribir=True)
            if not self._aviso_formulas_dado and self._contiene_formulas(libro):
                self._aviso_formulas_dado = True
                nombre_archivo = os.path.basename(self._ruta)
                print(f"⚠️  {nombre_archivo} contiene FÓRMULAS: "
                      "Python no calcula sus resultados.")
                print("💡 Tras guardar desde Python, abre y guarda el archivo "
                      "en Excel para que importar() vuelva a leer los "
                      "valores calculados de esas fórmulas.")
            if nombre in libro.sheetnames:
                if not sobrescribir:
                    return None
                indice = libro.sheetnames.index(nombre)
                libro.remove(libro[nombre])
                hoja = libro.create_sheet(nombre, indice)
                accion = "actualizada"
            else:
                hoja = libro.create_sheet(nombre)
                accion = "creada"

        for i, fila in enumerate(datos, start=1):
            for j, valor in enumerate(fila, start=1):
                hoja.cell(row=i, column=j, value=valor)

        libro.save(self._ruta)
        self._nuevo = False
        return accion


def _conectar(fuente, para_escribir=False):
    """Abre la fuente indicada y devuelve el backend adecuado (o None con mensajes)."""
    if _es_excel(fuente):
        try:
            return _BackendExcel(fuente, crear_si_no_existe=para_escribir)
        except FileNotFoundError:
            print(f"❌ No existe el archivo Excel: {fuente}")
            print("💡 Verifica la ruta, o usa exportar() para crearlo con tus variables")
            return None
        except ImportError as e:
            print(f"❌ {e}")
            return None

    # Las fuentes de Google (nombre o enlace) necesitan configurar()
    if not _verificar_configuracion():
        return None
    try:
        return _BackendGoogle(fuente)
    except Exception:
        print(f"❌ No se pudo abrir '{fuente}'")
        print("💡 Verifica que el archivo existe y tienes permisos")
        if not _es_enlace(fuente):
            _sugerir_sheets_disponibles()
        print("📝 Para cambiar de fuente: cambiar_sheet('nombre, enlace o ruta .xlsx')")
        return None


# ============================================================
# Clase auxiliar
# ============================================================

class MatrixLoader:
    """Clase auxiliar para cargar y gestionar matrices."""

    def __init__(self):
        self._datos = {}

    def __getattr__(self, name):
        if name in self._datos:
            return self._datos[name]
        raise AttributeError(f"No se encontró la matriz '{name}'")

    def __repr__(self):
        return f"MatrixLoader con matrices: {list(self._datos.keys())}"

    def listar_matrices(self):
        """Muestra todas las matrices disponibles."""
        if not self._datos:
            print("⚠️  No hay matrices cargadas")
            return

        print("📊 Matrices cargadas:")
        for nombre, matriz in self._datos.items():
            if isinstance(matriz, np.ndarray):
                print(f"  • {nombre}: {matriz.shape}")
            else:
                print(f"  • {nombre}: {matriz} (escalar)")


# ============================================================
# Configuración
# ============================================================

def configurar(sheet=None):
    """
    🔧 Configuración inicial del sistema.

    Autentica con Google y establece la conexión con Google Sheets.
    En modo Excel local (.xlsx) NO se necesita cuenta de Google.
    Ejecutar UNA VEZ al inicio de cada sesión.

    Args:
        sheet (str, optional): Fuente de trabajo. Acepta tres formas:
            configurar(sheet='prope2026')                  # Google Sheet por nombre
            configurar(sheet='https://docs.google.com/..') # Google Sheet por enlace
            configurar(sheet='C:/datos/matrices.xlsx')     # Excel local
            Si no se indica, se usa la actual ('matrices' por defecto).

    Returns:
        bool: True si la configuración fue exitosa
    """
    global gc, spreadsheet_name

    if sheet:
        spreadsheet_name = sheet

    # Modo Excel local: no necesita Google
    if _es_excel(spreadsheet_name):
        print(f"📊 Modo Excel local: {spreadsheet_name}")
        if os.path.exists(spreadsheet_name):
            print("✅ Archivo encontrado — no se necesita cuenta de Google")
        else:
            print("⚠️  El archivo aún no existe: exportar() lo creará")
        print("📖 Usa ayuda() para ver todos los comandos disponibles")
        return True

    try:
        # Verificar si estamos en Google Colab
        try:
            from google.colab import auth
            auth.authenticate_user()
            print("✅ Autenticación de Google Colab completada")
        except ImportError:
            print("ℹ️  Configurando autenticación local...")

        # Establecer conexión con Google Sheets
        creds, _ = default()
        gc = gspread.authorize(creds)

        print("✅ Conexión con Google Sheets establecida")
        if _es_enlace(spreadsheet_name):
            print("📋 Fuente de trabajo: enlace de Google Sheets")
        else:
            print(f"📋 Sheet de trabajo: '{spreadsheet_name}'")
        print("💡 Para cambiar de fuente: cambiar_sheet('nombre, enlace o ruta .xlsx')")
        print("📖 Usa ayuda() para ver todos los comandos disponibles")

        return True

    except Exception as e:
        print(f"❌ Error en configuración: {e}")
        print("💡 Soluciones:")
        print("   • En Google Colab: Reinicia runtime y vuelve a intentar")
        print("   • Localmente: Configura credenciales de Google Cloud")
        print("   • Para trabajar sin Google: configurar(sheet='mi_archivo.xlsx')")
        return False


def _sugerir_sheets_disponibles(max_mostrar=10):
    """Muestra los Google Sheets a los que la cuenta tiene acceso."""
    if gc is None:
        return
    try:
        titulos = [s.title for s in gc.openall()]
        if titulos:
            extra = f" (y {len(titulos) - max_mostrar} más)" if len(titulos) > max_mostrar else ""
            print(f"💡 Sheets disponibles en tu cuenta: {', '.join(titulos[:max_mostrar])}{extra}")
    except Exception:
        pass


def cambiar_sheet(fuente):
    """
    📝 Cambiar la fuente de trabajo.

    Acepta tres formas:
        cambiar_sheet('prope2026')                    # Google Sheet por nombre
        cambiar_sheet('https://docs.google.com/...')  # Google Sheet por enlace
        cambiar_sheet('C:/datos/matrices.xlsx')       # Excel local

    A partir de aquí, importar(), workspace() y exportar()
    usarán esta fuente.

    Args:
        fuente (str): Nombre, enlace o ruta .xlsx
    """
    global spreadsheet_name
    spreadsheet_name = fuente

    # Modo Excel local
    if _es_excel(fuente):
        if os.path.exists(fuente):
            print("✅ Archivo Excel encontrado")
        else:
            print("⚠️  El archivo aún no existe: exportar() lo creará")
        print(f"📊 Ahora trabajando en modo Excel local: {fuente}")
        return

    # Fuentes de Google: verificar que se puede abrir
    if gc is None:
        print("🔧 Recuerda ejecutar configurar() para conectar con Google")
    else:
        try:
            if _es_enlace(fuente):
                titulo = gc.open_by_url(fuente).title
            else:
                titulo = gc.open(fuente).title
            print(f"✅ Archivo '{titulo}' encontrado")
        except Exception:
            print("⚠️  No pude abrir la fuente con tu cuenta.")
            print("   Verifica que el nombre o enlace sea EXACTO y que tengas acceso.")
            if not _es_enlace(fuente):
                _sugerir_sheets_disponibles()

    if _es_enlace(fuente):
        print("📋 Ahora trabajando con el enlace de Google Sheets indicado")
    else:
        print(f"📋 Ahora trabajando con el sheet: '{fuente}'")


def _verificar_configuracion():
    """Verifica que el sistema esté configurado (solo necesario para Google)."""
    if gc is None:
        print("❌ Sistema no configurado.")
        print("🔧 Ejecuta: configurar()")
        print("📊 O trabaja sin Google: cambiar_sheet('mi_archivo.xlsx')")
        print("📦 Instalado desde: pip install algebra-lineal-sheets")
        return False
    return True


# ============================================================
# Utilidades internas
# ============================================================

def _limpiar_nombre_variable(nombre):
    """Convierte nombre de pestaña en variable válida de Python."""
    if not nombre:
        return "variable_sin_nombre"

    nombre_limpio = re.sub(r'[^a-zA-Z0-9_]', '_', nombre)
    if nombre_limpio and nombre_limpio[0].isdigit():
        nombre_limpio = f"var_{nombre_limpio}"
    return nombre_limpio or "variable_sin_nombre"


class ErrorDeDatos(ValueError):
    """Datos inválidos en una pestaña; el mensaje viene en español, listo para mostrar."""


def _coordenada_celda(fila, columna):
    """Índices (0, 0) → coordenada 'A1' estilo hoja de cálculo."""
    letras = ''
    c = columna
    while True:
        letras = chr(ord('A') + c % 26) + letras
        c = c // 26 - 1
        if c < 0:
            break
    return f"{letras}{fila + 1}"


def _normalizar_celda(celda):
    """Normaliza una celda cruda.

    Returns:
        None si está vacía, float si es numérica (acepta punto o coma decimal,
        p. ej. '3.5', '3,5' o '1.234,56'), o el texto original si no es numérica.
    """
    if celda is None:
        return None
    if isinstance(celda, bool):
        return float(celda)
    if isinstance(celda, (int, float, np.number)):
        return float(celda)

    texto = str(celda).strip()
    if texto == '':
        return None
    try:
        return float(texto)
    except ValueError:
        pass

    # Formato regional: '1.234,56' (miles con punto) o '3,5' (coma decimal)
    candidato = texto.replace(' ', '')
    if re.fullmatch(r'[+-]?\d{1,3}(\.\d{3})+(,\d+)?', candidato):
        candidato = candidato.replace('.', '').replace(',', '.')
    else:
        candidato = candidato.replace(',', '.')
    try:
        return float(candidato)
    except ValueError:
        return texto


def _limpiar_y_validar(nombre_pestana, datos_crudos):
    """Convierte los datos crudos de una pestaña en lista de filas de floats.

    - Acepta números nativos y textos numéricos con punto o coma decimal.
    - Elimina filas y columnas COMPLETAMENTE vacías (separadores legítimos).
    - Celdas de texto no numérico o huecos intermedios → ErrorDeDatos con
      coordenadas de hoja de cálculo, en vez de corromper la matriz.

    Returns:
        list[list[float]]: vacía si la pestaña no tiene datos.
    """
    if not datos_crudos:
        return []

    ancho = max((len(fila) for fila in datos_crudos), default=0)
    rejilla = []
    for fila in datos_crudos:
        fila_norm = [_normalizar_celda(celda) for celda in fila]
        fila_norm += [None] * (ancho - len(fila_norm))
        rejilla.append(fila_norm)

    filas_utiles = [i for i, fila in enumerate(rejilla)
                    if any(celda is not None for celda in fila)]
    cols_utiles = [j for j in range(ancho)
                   if any(rejilla[i][j] is not None for i in filas_utiles)]
    if not filas_utiles:
        return []

    textos = []
    huecos = []
    matriz = []
    for i in filas_utiles:
        fila_limpia = []
        for j in cols_utiles:
            celda = rejilla[i][j]
            if celda is None:
                huecos.append(_coordenada_celda(i, j))
            elif isinstance(celda, str):
                textos.append((_coordenada_celda(i, j), celda))
            else:
                fila_limpia.append(celda)
        matriz.append(fila_limpia)

    if textos or huecos:
        lineas = [f"❌ '{nombre_pestana}': la pestaña tiene datos "
                  "que no son numéricos"]
        if textos:
            detalle = ', '.join(f"{coord} ('{valor}')" for coord, valor in textos[:5])
            if len(textos) > 5:
                detalle += f" y {len(textos) - 5} más"
            lineas.append(f"   📝 Texto en: {detalle}")
            primera = filas_utiles[0]
            if all(isinstance(rejilla[primera][j], str) for j in cols_utiles):
                lineas.append("   💡 La primera fila parece de ENCABEZADOS: bórrala, "
                              "la matriz debe contener solo números")
        if huecos:
            detalle = ', '.join(huecos[:5])
            if len(huecos) > 5:
                detalle += f" y {len(huecos) - 5} más"
            lineas.append(f"   ⬜ Celdas vacías en: {detalle}")
            lineas.append("   💡 Completa los datos o elimina la fila/columna incompleta")
        raise ErrorDeDatos('\n'.join(lineas))

    return matriz


def _obtener_dimensiones_y_tipo(valor):
    """Obtiene las dimensiones y tipo de una variable de forma segura."""
    if isinstance(valor, (int, float, np.number)):
        return (1, 1), "📊 Escalar"
    elif isinstance(valor, np.ndarray):
        if valor.ndim == 0:
            return (1, 1), "📊 Escalar"
        elif valor.ndim == 1:
            return (len(valor), 1), "📉 Vector columna"
        elif valor.ndim == 2:
            if valor.shape[0] == 1:
                return valor.shape, "📈 Vector fila"
            elif valor.shape[1] == 1:
                return valor.shape, "📉 Vector columna"
            else:
                return valor.shape, "📋 Matriz"
        else:
            return valor.shape, "❓ Array ND"
    else:
        return None, "❌ Tipo no soportado"


def _convertir_a_valor(datos_limpios):
    """Convierte datos ya limpios (de _limpiar_y_validar) en escalar/vector/matriz.

    Returns:
        (valor, tipo_str)
    """
    array_np = np.array(datos_limpios, dtype=float)

    if array_np.size == 1:
        return array_np.item(), "escalar"
    elif array_np.shape[0] == 1 and array_np.shape[1] > 1:
        return array_np, "vector fila"
    elif array_np.shape[0] > 1 and array_np.shape[1] == 1:
        return array_np, "vector columna"
    else:
        return array_np, "matriz"


# ============================================================
# Funciones principales
# ============================================================

def workspace(sheet_name=None):
    """
    🏢 Muestra todas las matrices disponibles en la fuente activa.

    Args:
        sheet_name (str, optional): Fuente puntual (nombre, enlace o .xlsx).
            Si None, usa la configurada.

    Returns:
        dict: Información del workspace
    """
    nombre_sheet = sheet_name or spreadsheet_name
    backend = _conectar(nombre_sheet)
    if backend is None:
        return {}

    print(f"🏢 WORKSPACE — {backend.descripcion}")
    print("=" * 75)
    print(f"{'#':<3} {'NOMBRE':<20} {'DIMENSIONES':<12} {'TIPO':<15}")
    print("-" * 75)

    workspace_info = {}

    for i, nombre in enumerate(backend.nombres_pestanas(), 1):
        try:
            data = backend.leer(nombre)
            datos_limpios = _limpiar_y_validar(nombre, data)

            if not datos_limpios:
                print(f"{i:<3} {nombre:<20} {'VACÍA':<12} {'⚠️  Vacía':<15}")
                workspace_info[nombre] = {'tipo': 'vacía', 'dimensiones': None}
                continue

            # Dimensiones reales en la fuente
            filas_sheets = len(datos_limpios)
            cols_sheets = len(datos_limpios[0])

            if filas_sheets == 1 and cols_sheets == 1:
                tipo = "📊 Escalar"
                dimensiones_str = "1×1"
                workspace_info[nombre] = {
                    'tipo': 'escalar', 'dimensiones': (1, 1)}

            elif filas_sheets == 1 and cols_sheets > 1:
                tipo = "📈 Vector fila"
                dimensiones_str = f"1×{cols_sheets}"
                workspace_info[nombre] = {
                    'tipo': 'vector_fila', 'dimensiones': (1, cols_sheets)}

            elif filas_sheets > 1 and cols_sheets == 1:
                tipo = "📉 Vector columna"
                dimensiones_str = f"{filas_sheets}×1"
                workspace_info[nombre] = {
                    'tipo': 'vector_columna', 'dimensiones': (filas_sheets, 1)}

            else:
                tipo = "📋 Matriz"
                dimensiones_str = f"{filas_sheets}×{cols_sheets}"
                workspace_info[nombre] = {
                    'tipo': 'matriz', 'dimensiones': (filas_sheets, cols_sheets)}

            print(f"{i:<3} {nombre:<20} {dimensiones_str:<12} {tipo:<15}")

        except ErrorDeDatos:
            print(f"{i:<3} {nombre:<20} {'REVISAR':<12} {'⚠️  Datos no numéricos':<15}")
            workspace_info[nombre] = {'tipo': 'datos_invalidos', 'dimensiones': None}

        except Exception:
            print(f"{i:<3} {nombre:<20} {'ERROR':<12} {'❌ Error':<15}")
            workspace_info[nombre] = {'tipo': 'error', 'dimensiones': None}

    print("=" * 75)
    print("💡 Usa importar('nombre') para traer matrices específicas")

    return workspace_info


def importar(*nombres_matrices, sheet_name=None):
    """
    📥 Importa matrices desde la fuente activa a variables globales.

    Ejemplos:
        importar()           # Importa TODAS las matrices
        importar('A')        # Importa solo A
        importar('A', 'B')   # Importa A y B
        importar('A', sheet_name='otro_archivo')   # Desde otra fuente puntual

    Args:
        *nombres_matrices: Nombres de matrices a importar
        sheet_name (str, optional): Fuente puntual (nombre, enlace o .xlsx)

    Returns:
        dict: Diccionario con las variables importadas
    """
    nombre_sheet = sheet_name or spreadsheet_name
    backend = _conectar(nombre_sheet)
    if backend is None:
        return {}

    if not nombres_matrices:
        print(f"🔄 Importando TODAS las matrices — {backend.descripcion}")
        return _cargar_todas_matrices(backend)

    print(f"🔄 Importando desde {backend.descripcion}: {', '.join(nombres_matrices)}")
    print("=" * 50)

    datos_cargados = {}
    disponibles = backend.nombres_pestanas()

    for nombre_solicitado in nombres_matrices:
        # Búsqueda exacta y luego case-insensitive
        pestana_encontrada = None
        if nombre_solicitado in disponibles:
            pestana_encontrada = nombre_solicitado
        else:
            for nombre_pestana in disponibles:
                if nombre_pestana.lower() == nombre_solicitado.lower():
                    pestana_encontrada = nombre_pestana
                    break

        if pestana_encontrada:
            try:
                data = backend.leer(pestana_encontrada)
                datos_limpios = _limpiar_y_validar(pestana_encontrada, data)

                if not datos_limpios:
                    print(f"⚠️  '{nombre_solicitado}' está vacía")
                    continue

                valor, tipo = _convertir_a_valor(datos_limpios)

                nombre_variable = _limpiar_nombre_variable(nombre_solicitado)
                datos_cargados[nombre_variable] = valor

                if isinstance(valor, (int, float)):
                    dimensiones_str = "1×1"
                else:
                    dimensiones_str = f"{valor.shape}"

                print(f"✅ {nombre_solicitado} → {tipo} {dimensiones_str}")

            except ErrorDeDatos as e:
                print(e)
            except Exception as e:
                print(f"❌ Error: {nombre_solicitado} - {e}")
        else:
            print(f"❌ No encontré la pestaña '{nombre_solicitado}' — {backend.descripcion}")
            print(f"💡 Pestañas disponibles: {', '.join(disponibles)}")
            print(f"ℹ️  ¿'{nombre_solicitado}' es un ARCHIVO y no una pestaña?")
            print(f"   Entonces usa: cambiar_sheet('{nombre_solicitado}') y luego importar()")

    if datos_cargados:
        # Crear variables en el frame del llamador (notebook del usuario)
        frame = inspect.currentframe().f_back
        if frame:
            frame.f_globals.update(datos_cargados)

        print(
            f"\n🎯 ¡Listo! Variables disponibles: {', '.join(datos_cargados.keys())}")

    print("=" * 50)
    return datos_cargados


def _cargar_todas_matrices(backend):
    """Carga todas las matrices de la fuente."""
    datos_cargados = {}

    print("=" * 50)
    for nombre_hoja in backend.nombres_pestanas():
        nombre_variable = _limpiar_nombre_variable(nombre_hoja)

        try:
            data = backend.leer(nombre_hoja)
            datos_limpios = _limpiar_y_validar(nombre_hoja, data)

            if not datos_limpios:
                print(f"⚠️  '{nombre_hoja}' está vacía")
                continue

            valor, tipo = _convertir_a_valor(datos_limpios)
            datos_cargados[nombre_variable] = valor

            if isinstance(valor, (int, float)):
                dimensiones_str = "1×1"
            else:
                dimensiones_str = f"{valor.shape}"

            print(f"✅ {nombre_hoja} → {tipo} {dimensiones_str}")

        except ErrorDeDatos as e:
            print(e)
        except Exception as e:
            print(f"❌ Error en {nombre_hoja}: {e}")

    if datos_cargados:
        # Crear variables en el frame del llamador (dos niveles arriba)
        frame = inspect.currentframe().f_back.f_back
        if frame:
            frame.f_globals.update(datos_cargados)

        print(f"\n🎯 Variables creadas: {', '.join(datos_cargados.keys())}")

    print("=" * 50)
    return datos_cargados


def exportar(*nombres_variables, sheet_name=None, sobrescribir=True):
    """
    📤 Exporta variables a la fuente activa.

    En modo Excel local, si el archivo no existe, se crea automáticamente.

    Ejemplos:
        exportar()              # Exporta TODAS las variables numpy
        exportar('C')           # Exporta solo C
        exportar('C', 'Ainv')   # Exporta C y Ainv
        exportar('C', sheet_name='resultados.xlsx')   # A otra fuente puntual

    Args:
        *nombres_variables: Nombres de variables a exportar
        sheet_name (str, optional): Fuente puntual (nombre, enlace o .xlsx)
        sobrescribir (bool): Si sobrescribir pestañas existentes

    Returns:
        list: Lista de variables exportadas exitosamente
    """
    nombre_sheet = sheet_name or spreadsheet_name
    backend = _conectar(nombre_sheet, para_escribir=True)
    if backend is None:
        return []

    # Obtener variables del frame del llamador
    frame = inspect.currentframe().f_back
    todas_variables = frame.f_globals if frame else {}

    # Detectar variables exportables
    variables_disponibles = {}
    for nombre, valor in todas_variables.items():
        if (isinstance(valor, (np.ndarray, np.number, int, float)) and
            not nombre.startswith('_') and
                nombre not in ['np', 'numpy', 'gspread', 'gc', 'creds', 'spreadsheet']):
            variables_disponibles[nombre] = valor

    if not nombres_variables:
        variables_a_exportar = variables_disponibles
        print("📤 Exportando TODAS las variables numpy...")
    else:
        variables_a_exportar = {}
        variables_no_encontradas = []

        for nombre in nombres_variables:
            if nombre in variables_disponibles:
                variables_a_exportar[nombre] = variables_disponibles[nombre]
            else:
                variables_no_encontradas.append(nombre)

        if variables_no_encontradas:
            print(
                f"⚠️  Variables no encontradas: {', '.join(variables_no_encontradas)}")
            print(
                f"💡 Variables disponibles: {list(variables_disponibles.keys())}")

        if not variables_a_exportar:
            print("❌ No hay variables válidas para exportar")
            return []

    print("=" * 50)
    exportadas = []

    for nombre_var, valor in variables_a_exportar.items():
        try:
            # Preparar datos para la fuente
            datos = _preparar_datos_para_sheets(valor)
            if datos is None:
                print(f"⚠️  {nombre_var}: Tipo no soportado")
                continue

            if _contiene_no_finitos(datos):
                print(f"❌ '{nombre_var}' contiene NaN o infinito — no se puede exportar")
                print("💡 Revisa divisiones entre cero o matrices singulares (determinante 0)")
                continue

            accion = backend.escribir(nombre_var, datos, sobrescribir=sobrescribir)
            if accion is None:
                print(f"⚠️  {nombre_var}: Ya existe (sobrescribir=False)")
                continue

            exportadas.append(nombre_var)
            print(f"✅ {nombre_var} → {accion}")

        except Exception as e:
            print(f"❌ Error exportando {nombre_var}: {e}")

    print("=" * 50)
    if exportadas:
        print(f"🎯 ¡Exportadas exitosamente!: {', '.join(exportadas)}")
    print(f"🔗 Revisa tu archivo — {backend.descripcion}")

    return exportadas


def _contiene_no_finitos(datos):
    """True si alguna celda es NaN o infinito (romperían la fuente de datos)."""
    return any(not np.isfinite(celda) for fila in datos for celda in fila)


def _preparar_datos_para_sheets(valor):
    """Convierte una variable en formato adecuado para la fuente de datos."""
    if isinstance(valor, (int, float, np.number)):
        return [[float(valor)]]
    elif isinstance(valor, np.ndarray):
        if valor.ndim == 0:
            # Escalar en array
            return [[float(valor.item())]]
        elif valor.ndim == 1:
            # Vector 1D: convertir a columna por defecto
            return [[float(x)] for x in valor]
        elif valor.ndim == 2:
            if valor.shape[0] == 1:
                # VECTOR FILA (1, n): exportar horizontalmente
                return [valor[0].tolist()]
            elif valor.shape[1] == 1:
                # VECTOR COLUMNA (n, 1): exportar verticalmente
                return [[float(valor[i, 0])] for i in range(valor.shape[0])]
            else:
                # MATRIZ: exportar normalmente
                return [[float(x) for x in fila] for fila in valor]
    return None


def listar_variables_exportables():
    """
    🔍 Lista todas las variables que se pueden exportar.

    Returns:
        dict: Variables exportables
    """
    frame = inspect.currentframe().f_back
    todas_variables = frame.f_globals if frame else {}

    print("🔍 VARIABLES EXPORTABLES:")
    print("=" * 50)

    variables_numpy = {}
    for nombre, valor in todas_variables.items():
        if (isinstance(valor, (np.ndarray, np.number, int, float)) and
            not nombre.startswith('_') and
                nombre not in ['np', 'numpy', 'gspread', 'gc', 'creds', 'spreadsheet']):
            variables_numpy[nombre] = valor

    if variables_numpy:
        print(f"{'NOMBRE':<15} {'DIMENSIONES':<12} {'TIPO':<15}")
        print("-" * 50)

        for nombre, valor in variables_numpy.items():
            dimensiones, tipo = _obtener_dimensiones_y_tipo(valor)

            if dimensiones:
                dimensiones_str = f"{dimensiones[0]}×{dimensiones[1]}"
            else:
                dimensiones_str = "N/A"

            print(f"{nombre:<15} {dimensiones_str:<12} {tipo:<15}")

        print("=" * 50)
        print(f"📊 Total: {len(variables_numpy)} variables exportables")
        print(
            f"💡 Uso: exportar() para todas, o exportar('{list(variables_numpy.keys())[0]}') para específicas")
    else:
        print("⚠️  No hay variables numpy disponibles para exportar")
        print("💡 Crea algunas variables primero:")
        print("   A = np.array([[1, 2], [3, 4]])")
        print("   C = A @ B")

    return variables_numpy


def ayuda():
    """
    📖 Muestra la ayuda completa del sistema.
    """
    print("📚 ALGEBRA LINEAL - GUÍA COMPLETA")
    print("=" * 60)
    print("📦 Instalación:")
    print("   pip install algebra-lineal-sheets")
    print()
    print("🔧 CONFIGURACIÓN:")
    print("   configurar()                 # Configurar una vez al inicio")
    print("   configurar(sheet='prope2026')             # Google Sheet por nombre")
    print("   configurar(sheet='https://docs.goo...')   # Google Sheet por enlace")
    print("   configurar(sheet='C:/datos/notas.xlsx')   # Excel local (sin Google)")
    print("   cambiar_sheet('fuente')      # Cambiar de archivo en cualquier momento")
    print()
    print("🏢 VER CONTENIDO:")
    print("   workspace()                  # Ver matrices en la fuente activa")
    print()
    print("📥 IMPORTAR:")
    print("   importar()                   # Importar todas las matrices")
    print("   importar('A')                # Importar solo A")
    print("   importar('A', 'B', 'v')      # Importar A, B y v")
    print("   importar('A', sheet_name='otro_archivo')  # Desde otra fuente")
    print()
    print("🧮 OPERACIONES (después de importar):")
    print("   C = A @ B                    # Multiplicación matricial")
    print("   suma = A + B                 # Suma")
    print("   Ainv = np.linalg.inv(A)      # Matriz inversa")
    print("   det_A = np.linalg.det(A)     # Determinante")
    print("   norma = np.linalg.norm(v)    # Norma de vector")
    print()
    print("📤 EXPORTAR:")
    print("   exportar()                   # Exportar todas las variables")
    print("   exportar('C')                # Exportar solo C")
    print("   exportar('C', 'Ainv')        # Exportar C y Ainv")
    print("   exportar('C', sheet_name='resultados.xlsx')  # A un Excel nuevo")
    print()
    print("🔍 UTILIDADES:")
    print("   listar_variables_exportables() # Ver qué se puede exportar")
    print("   version()                    # Información del paquete")
    print("   ayuda()                      # Esta ayuda")
    print("=" * 60)


def version():
    """Muestra información del paquete."""
    try:
        from . import __version__, __author__, __description__, __url__
        print(f"📦 ALGEBRA LINEAL v{__version__}")
        print(f"👨‍🏫 Autor: {__author__}")
        print(f"📚 {__description__}")
        print(f"🔗 PyPI: {__url__}")
    except ImportError:
        print("📦 ALGEBRA LINEAL")
        print("👨‍🏫 Sistema de álgebra lineal con Google Sheets y Excel")
    print("🛠️  Instalación: pip install algebra-lineal-sheets")
