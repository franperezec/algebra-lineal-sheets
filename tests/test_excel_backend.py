# -*- coding: utf-8 -*-
"""Tests del backend Excel local y del round-trip por la API pública."""
import numpy as np
import openpyxl

import algebra_lineal.core as core
from algebra_lineal.core import _BackendExcel, _limpiar_y_validar


def test_escribir_celdas_numericas_no_texto(tmp_path):
    """Las celdas exportadas son números nativos: Excel les aplica el formato
    regional del usuario (coma o punto), nunca se escriben como texto."""
    ruta = str(tmp_path / 'm.xlsx')
    backend = _BackendExcel(ruta, crear_si_no_existe=True)
    backend.escribir('A', [[1.5, 2.0], [3.25, 4.0]])
    backend.guardar()

    libro = openpyxl.load_workbook(ruta)
    celdas = [c.value for fila in libro['A'].iter_rows() for c in fila]
    assert all(isinstance(v, (int, float)) for v in celdas)
    assert celdas == [1.5, 2.0, 3.25, 4.0]


def test_round_trip_con_decimales(tmp_path):
    # Regresión: antes leer() pasaba todo por str() y reparseaba
    ruta = str(tmp_path / 'm.xlsx')
    backend = _BackendExcel(ruta, crear_si_no_existe=True)
    original = [[1.5, -2.75], [0.001, 1234.56]]
    backend.escribir('A', original)
    backend.guardar()

    leidos = _limpiar_y_validar('A', backend.leer('A'))
    assert leidos == original


def test_sobrescribir_false_no_toca_la_pestana(tmp_path):
    ruta = str(tmp_path / 'm.xlsx')
    backend = _BackendExcel(ruta, crear_si_no_existe=True)
    backend.escribir('A', [[1.0]])
    backend.guardar()
    assert backend.escribir('A', [[2.0]], sobrescribir=False) is None
    assert _limpiar_y_validar('A', backend.leer('A')) == [[1.0]]


def test_aviso_de_formulas_al_escribir(tmp_path, capsys):
    ruta = str(tmp_path / 'm.xlsx')
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = 'datos'
    hoja['A1'] = 1
    hoja['A2'] = '=A1*2'
    libro.save(ruta)

    backend = _BackendExcel(ruta)
    backend.escribir('B', [[9.0]])
    salida = capsys.readouterr().out
    assert 'FÓRMULAS' in salida


def test_exportar_importar_api_publica(tmp_path):
    """Round-trip completo por la API pública con fuente puntual."""
    ruta = str(tmp_path / 'pub.xlsx')
    globals()['M_TEST_RT'] = np.array([[1.5, 2.5], [3.5, 4.5]])
    try:
        exportadas = core.exportar('M_TEST_RT', sheet_name=ruta)
        assert exportadas == ['M_TEST_RT']

        resultado = core.importar('M_TEST_RT', sheet_name=ruta)
        assert np.array_equal(resultado['M_TEST_RT'], globals()['M_TEST_RT'])
    finally:
        globals().pop('M_TEST_RT', None)


def test_exportar_rechaza_nan(tmp_path, capsys):
    ruta = str(tmp_path / 'nan.xlsx')
    globals()['M_TEST_NAN'] = np.array([[1.0, np.nan]])
    try:
        exportadas = core.exportar('M_TEST_NAN', sheet_name=ruta)
        assert exportadas == []
        assert 'NaN' in capsys.readouterr().out
    finally:
        globals().pop('M_TEST_NAN', None)


def _contar_cargas_y_guardados(monkeypatch):
    """Parcha openpyxl para contar load_workbook() y Workbook.save()."""
    contador = {'cargas': 0, 'guardados': 0}

    cargar_original = openpyxl.load_workbook

    def cargar_contando(*args, **kwargs):
        contador['cargas'] += 1
        return cargar_original(*args, **kwargs)

    guardar_original = openpyxl.workbook.workbook.Workbook.save

    def guardar_contando(self, *args, **kwargs):
        contador['guardados'] += 1
        return guardar_original(self, *args, **kwargs)

    monkeypatch.setattr(openpyxl, 'load_workbook', cargar_contando)
    monkeypatch.setattr(openpyxl.workbook.workbook.Workbook, 'save',
                        guardar_contando)
    return contador


def test_exportar_carga_y_guarda_una_sola_vez(tmp_path, monkeypatch):
    """Exportar N variables debe cargar y guardar el libro UNA vez, no N."""
    ruta = str(tmp_path / 'lote.xlsx')
    libro = openpyxl.Workbook()
    libro.active.title = 'previa'
    libro.active['A1'] = 1.0
    libro.save(ruta)

    globals()['V1_LOTE'] = np.array([[1.0]])
    globals()['V2_LOTE'] = np.array([[2.0]])
    globals()['V3_LOTE'] = np.array([[3.0]])
    contador = _contar_cargas_y_guardados(monkeypatch)
    try:
        exportadas = core.exportar('V1_LOTE', 'V2_LOTE', 'V3_LOTE',
                                   sheet_name=ruta)
        assert exportadas == ['V1_LOTE', 'V2_LOTE', 'V3_LOTE']
        assert contador['cargas'] == 1
        assert contador['guardados'] == 1
    finally:
        for nombre in ('V1_LOTE', 'V2_LOTE', 'V3_LOTE'):
            globals().pop(nombre, None)

    final = openpyxl.load_workbook(ruta)
    assert {'V1_LOTE', 'V2_LOTE', 'V3_LOTE'} <= set(final.sheetnames)


def test_importar_carga_el_libro_una_sola_vez(tmp_path, monkeypatch):
    """importar() (nombres_pestanas + leer por pestaña) carga el libro UNA vez."""
    ruta = str(tmp_path / 'lectura.xlsx')
    libro = openpyxl.Workbook()
    libro.active.title = 'M1'
    libro.active['A1'] = 1.0
    for nombre in ('M2', 'M3'):
        libro.create_sheet(nombre)['A1'] = 2.0
    libro.save(ruta)

    contador = _contar_cargas_y_guardados(monkeypatch)
    try:
        resultado = core.importar(sheet_name=ruta)
    finally:
        for nombre in ('M1', 'M2', 'M3'):
            globals().pop(nombre, None)
    assert set(resultado) == {'M1', 'M2', 'M3'}
    assert contador['cargas'] == 1


def test_guardar_fallido_avisa_en_espanol(tmp_path, monkeypatch, capsys):
    """Si save() falla (p. ej. archivo abierto en Excel), exportar() devuelve []."""
    ruta = str(tmp_path / 'bloqueado.xlsx')

    def guardar_fallando(self, *args, **kwargs):
        raise PermissionError('Permission denied')

    monkeypatch.setattr(openpyxl.workbook.workbook.Workbook, 'save',
                        guardar_fallando)
    globals()['V_BLOQ'] = np.array([[1.0]])
    try:
        exportadas = core.exportar('V_BLOQ', sheet_name=ruta)
    finally:
        globals().pop('V_BLOQ', None)
    salida = capsys.readouterr().out
    assert exportadas == []
    assert 'No se pudo guardar' in salida
    assert '¡Exportadas exitosamente!' not in salida


def test_escribir_sin_guardar_no_toca_el_disco(tmp_path):
    """Contrato nuevo: escribir() queda en memoria hasta guardar()."""
    ruta = str(tmp_path / 'memoria.xlsx')
    libro = openpyxl.Workbook()
    libro.active.title = 'previa'
    libro.active['A1'] = 1.0
    libro.save(ruta)

    backend = _BackendExcel(ruta)
    backend.escribir('B', [[9.0]])
    assert 'B' not in openpyxl.load_workbook(ruta).sheetnames
    backend.guardar()
    assert 'B' in openpyxl.load_workbook(ruta).sheetnames


def test_archivo_nuevo_con_varias_variables(tmp_path):
    """El archivo nuevo se crea con las pestañas exportadas y sin 'Sheet' sobrante."""
    ruta = str(tmp_path / 'nuevo.xlsx')
    globals()['VA_NUEVO'] = np.array([[1.0]])
    globals()['VB_NUEVO'] = np.array([[2.0]])
    try:
        exportadas = core.exportar('VA_NUEVO', 'VB_NUEVO', sheet_name=ruta)
        assert exportadas == ['VA_NUEVO', 'VB_NUEVO']
    finally:
        globals().pop('VA_NUEVO', None)
        globals().pop('VB_NUEVO', None)
    assert openpyxl.load_workbook(ruta).sheetnames == ['VA_NUEVO', 'VB_NUEVO']


def test_guardar_invalida_cache_de_lectura(tmp_path):
    """Round-trip en la MISMA instancia: leer → escribir → guardar → leer."""
    ruta = str(tmp_path / 'cache.xlsx')
    libro = openpyxl.Workbook()
    libro.active.title = 'previa'
    libro.active['A1'] = 1.0
    libro.save(ruta)

    backend = _BackendExcel(ruta)
    assert _limpiar_y_validar('previa', backend.leer('previa')) == [[1.0]]
    backend.escribir('B', [[7.5]])
    backend.guardar()
    assert _limpiar_y_validar('B', backend.leer('B')) == [[7.5]]


def test_aviso_formulas_una_sola_vez(tmp_path, capsys):
    """Dos escribir() en la misma instancia → un solo aviso de fórmulas."""
    ruta = str(tmp_path / 'form.xlsx')
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = 'datos'
    hoja['A1'] = 1
    hoja['A2'] = '=A1*2'
    libro.save(ruta)

    backend = _BackendExcel(ruta)
    backend.escribir('B', [[9.0]])
    backend.escribir('C', [[8.0]])
    salida = capsys.readouterr().out
    assert salida.count('FÓRMULAS') == 1
